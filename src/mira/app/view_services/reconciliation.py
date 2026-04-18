# SPDX-License-Identifier: GPL-3.0-or-later
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.

"""Module documentation."""

from __future__ import annotations

import hashlib
import unicodedata
from collections import Counter
from dataclasses import dataclass, replace
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook

from mira.db.database import Database
from mira.transaction_kinds import is_balance_adjustment_transaction


def _normalize_header(value: object) -> str:
    """Return normalize header."""
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    folded = "".join(char for char in text if not unicodedata.combining(char))
    return folded.casefold()


def _parse_excel_date(value: object) -> str:
    """Return parse excel date."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw_value = str(value or "").strip()
    if not raw_value:
        raise ValueError("Date is required.")
    for parser in (datetime.fromisoformat,):
        try:
            return parser(raw_value).date().isoformat()
        except ValueError:
            continue
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw_value, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Invalid date: {raw_value!r}.")


def _parse_excel_amount(value: object, *, field_name: str) -> float | None:
    """Return parse excel amount."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw_value = str(value).strip()
    if not raw_value:
        return None
    try:
        return float(raw_value.replace(",", ""))
    except ValueError as exc:
        raise ValueError(f"Invalid {field_name} amount: {raw_value!r}.") from exc


def _fingerprint_external_row(*, tx_date: str, reference: str, description: str, amount: float) -> str:
    """Return fingerprint external row."""
    payload = "|".join(
        (
            tx_date,
            reference.strip().casefold(),
            description.strip().casefold(),
            f"{amount:.2f}",
        )
    )
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:20]


@dataclass(frozen=True)
class ReconciliationParams:
    """Represent the ReconciliationParams class."""

    account_id: int | None

    date_from: str
    date_to: str


@dataclass(frozen=True)
class ReconciliationPanelSummary:
    """Represent the ReconciliationPanelSummary class."""

    opening_balance: float

    total_income: float
    total_expense: float
    closing_balance: float
    currency: str
    opening_balance_editable: bool


@dataclass(frozen=True)
class ReconciliationPreviewIssue:
    """Represent the ReconciliationPreviewIssue class."""

    row_number: int

    raw_date: str
    reference: str
    description: str
    raw_income: str
    raw_expense: str
    error: str


@dataclass(frozen=True)
class ReconciliationExternalRow:
    """Represent the ReconciliationExternalRow class."""

    row_number: int

    date: str
    reference: str
    description: str
    amount: float
    external_item_key: str
    is_reconciled: bool = False
    suggested_transaction_ids: tuple[int, ...] = ()
    group_ids: tuple[str, ...] = ()


@dataclass(frozen=True)
class ReconciliationSystemRow:
    """Represent the ReconciliationSystemRow class."""

    transaction_id: int

    date: str
    description: str
    amount: float
    tx_type: str
    payment_method: str
    is_transfer: bool
    is_reconciled: bool
    reconciled_group_ids: tuple[str, ...]
    selectable: bool


@dataclass(frozen=True)
class ReconciliationImportPreview:
    """Represent the ReconciliationImportPreview class."""

    filepath: str

    valid_rows: tuple[ReconciliationExternalRow, ...]
    invalid_rows: tuple[ReconciliationPreviewIssue, ...]
    missing_columns: tuple[str, ...]

    @property
    def has_blocking_error(self) -> bool:
        """Return whether  blocking error."""
        return bool(self.missing_columns)


@dataclass(frozen=True)
class ReconciliationViewState:
    """Represent the ReconciliationViewState class."""

    params: ReconciliationParams

    accounts: tuple[dict[str, Any], ...]
    external_summary: ReconciliationPanelSummary
    system_summary: ReconciliationPanelSummary
    external_rows: tuple[ReconciliationExternalRow, ...]
    system_rows: tuple[ReconciliationSystemRow, ...]
    groups: tuple[dict[str, Any], ...]
    amount_difference: float


class ReconciliationViewService:
    """Represent the ReconciliationViewService class."""

    _REQUIRED_COLUMNS = {
        "fecha": "Fecha",
        "referencia": "Referencia",
        "descripcion": "Descripción",
        "ingreso": "Ingreso",
        "gastos": "Gastos",
    }

    def __init__(self, db: Database) -> None:
        """Initialize the ReconciliationViewService instance."""
        self._db = db

    def list_accounts(self) -> tuple[dict[str, Any], ...]:
        """Return list accounts."""
        return tuple(self._db.account.list())

    def parse_excel(self, filepath: str) -> ReconciliationImportPreview:
        """Return parse excel."""
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            assert worksheet is not None
            header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_cells is None:
                return ReconciliationImportPreview(
                    filepath=filepath,
                    valid_rows=(),
                    invalid_rows=(),
                    missing_columns=tuple(self._REQUIRED_COLUMNS.values()),
                )
            header_map = {_normalize_header(value): index for index, value in enumerate(header_cells)}
            missing_columns = tuple(
                display_name
                for normalized_name, display_name in self._REQUIRED_COLUMNS.items()
                if normalized_name not in header_map
            )
            if missing_columns:
                return ReconciliationImportPreview(
                    filepath=filepath,
                    valid_rows=(),
                    invalid_rows=(),
                    missing_columns=missing_columns,
                )

            valid_rows: list[ReconciliationExternalRow] = []
            invalid_rows: list[ReconciliationPreviewIssue] = []
            occurrence_counter: Counter[str] = Counter()
            for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                raw_date = row[header_map["fecha"]]
                raw_reference = row[header_map["referencia"]]
                raw_description = row[header_map["descripcion"]]
                raw_income = row[header_map["ingreso"]]
                raw_expense = row[header_map["gastos"]]
                try:
                    tx_date = _parse_excel_date(raw_date)
                    reference = str(raw_reference or "").strip()
                    description = str(raw_description or "").strip()
                    income_amount = _parse_excel_amount(raw_income, field_name="income")
                    expense_amount = _parse_excel_amount(raw_expense, field_name="expense")
                    if (income_amount is None and expense_amount is None) or (
                        income_amount is not None and expense_amount is not None
                    ):
                        raise ValueError("Each row must contain either Income or Expenses, but not both.")
                    amount = (
                        abs(float(income_amount)) if income_amount is not None else -abs(float(expense_amount or 0.0))
                    )
                    fingerprint = _fingerprint_external_row(
                        tx_date=tx_date,
                        reference=reference,
                        description=description,
                        amount=amount,
                    )
                    occurrence_counter[fingerprint] += 1
                    external_item_key = f"{fingerprint}:{occurrence_counter[fingerprint]}"
                    valid_rows.append(
                        ReconciliationExternalRow(
                            row_number=row_number,
                            date=tx_date,
                            reference=reference,
                            description=description,
                            amount=round(amount, 2),
                            external_item_key=external_item_key,
                        )
                    )
                except ValueError as exc:
                    invalid_rows.append(
                        ReconciliationPreviewIssue(
                            row_number=row_number,
                            raw_date=str(raw_date or ""),
                            reference=str(raw_reference or ""),
                            description=str(raw_description or ""),
                            raw_income=str(raw_income or ""),
                            raw_expense=str(raw_expense or ""),
                            error=str(exc),
                        )
                    )
            return ReconciliationImportPreview(
                filepath=filepath,
                valid_rows=tuple(valid_rows),
                invalid_rows=tuple(invalid_rows),
                missing_columns=(),
            )
        finally:
            workbook.close()

    @staticmethod
    def _suggestions_for_external_row(
        external_row: ReconciliationExternalRow,
        system_rows: tuple[ReconciliationSystemRow, ...],
    ) -> tuple[int, ...]:
        """Return suggestions for external row."""
        candidates: list[tuple[int, int]] = []
        external_day = date.fromisoformat(external_row.date)
        for system_row in system_rows:
            if not system_row.selectable:
                continue
            if round(system_row.amount, 2) != round(external_row.amount, 2):
                continue
            score = 10
            system_day = date.fromisoformat(system_row.date)
            day_distance = abs((system_day - external_day).days)
            if day_distance == 0:
                score += 4
            elif day_distance <= 2:
                score += 2
            haystack = f"{system_row.description} {external_row.reference} {external_row.description}".casefold()
            if external_row.reference and external_row.reference.casefold() in haystack:
                score += 2
            if external_row.description and external_row.description.casefold() in haystack:
                score += 1
            candidates.append((score, system_row.transaction_id))
        candidates.sort(key=lambda item: (-item[0], item[1]))
        if not candidates:
            return ()
        top_score = candidates[0][0]
        return tuple(transaction_id for score, transaction_id in candidates if score == top_score)

    def load_state(
        self,
        *,
        account_id: int,
        date_from: str,
        date_to: str,
        external_rows: tuple[ReconciliationExternalRow, ...] = (),
        external_opening_balance: float = 0.0,
    ) -> ReconciliationViewState:
        """Return load state."""
        account = self._db.account.get(account_id)
        if account is None:
            raise ValueError(f"Account {account_id} not found.")
        from_day = date.fromisoformat(date_from)
        prev_day = (from_day - timedelta(days=1)).isoformat()
        system_transactions = tuple(
            sorted(
                self._db.transaction.list(
                    limit=20_000,
                    account_id=account_id,
                    since_date=date_from,
                    until_date=date_to,
                ),
                key=lambda item: (str(item.get("date") or ""), int(item.get("id") or 0)),
            )
        )
        groups = tuple(self._db.reconciliation.list_groups(account_id=account_id, date_from=date_from, date_to=date_to))
        matches = tuple(
            self._db.reconciliation.list_matches(account_id=account_id, date_from=date_from, date_to=date_to)
        )
        groups_by_transaction: dict[int, set[str]] = {}
        groups_by_external_key: dict[str, set[str]] = {}
        for match in matches:
            groups_by_transaction.setdefault(int(match["system_transaction_id"]), set()).add(
                str(match["reconciliation_group_id"])
            )
            groups_by_external_key.setdefault(str(match["external_item_key"]), set()).add(
                str(match["reconciliation_group_id"])
            )

        system_rows = tuple(
            ReconciliationSystemRow(
                transaction_id=int(tx["id"]),
                date=str(tx["date"]),
                description=str(tx.get("description") or ""),
                amount=float(tx["amount"]) if str(tx.get("type") or "") == "income" else -float(tx["amount"]),
                tx_type=str(tx.get("type") or ""),
                payment_method=str(tx.get("payment_method") or ""),
                is_transfer=int(tx.get("is_transfer") or 0) == 1,
                is_reconciled=int(tx.get("is_reconciled") or 0) == 1,
                reconciled_group_ids=tuple(sorted(groups_by_transaction.get(int(tx["id"]), set()))),
                selectable=int(tx.get("is_transfer") or 0) == 0 and not is_balance_adjustment_transaction(tx),
            )
            for tx in system_transactions
        )

        enriched_external_rows = tuple(
            replace(
                row,
                is_reconciled=bool(groups_by_external_key.get(row.external_item_key)),
                suggested_transaction_ids=self._suggestions_for_external_row(row, system_rows),
                group_ids=tuple(sorted(groups_by_external_key.get(row.external_item_key, set()))),
            )
            for row in external_rows
        )

        system_income = sum(max(row.amount, 0.0) for row in system_rows)
        system_expense = abs(sum(min(row.amount, 0.0) for row in system_rows))
        balance_snapshot = self._db.account.balance_as_of(account_id, prev_day)
        system_opening_balance = float(balance_snapshot["balance_as_of"])
        system_summary = ReconciliationPanelSummary(
            opening_balance=system_opening_balance,
            total_income=round(system_income, 2),
            total_expense=round(system_expense, 2),
            closing_balance=round(system_opening_balance + system_income - system_expense, 2),
            currency=str(account.get("currency") or self._db.setting.get_default_currency()).strip().upper(),
            opening_balance_editable=False,
        )

        external_income = sum(max(row.amount, 0.0) for row in enriched_external_rows)
        external_expense = abs(sum(min(row.amount, 0.0) for row in enriched_external_rows))
        external_summary = ReconciliationPanelSummary(
            opening_balance=round(float(external_opening_balance), 2),
            total_income=round(external_income, 2),
            total_expense=round(external_expense, 2),
            closing_balance=round(float(external_opening_balance) + external_income - external_expense, 2),
            currency=str(account.get("currency") or self._db.setting.get_default_currency()).strip().upper(),
            opening_balance_editable=True,
        )

        selected_external_total = round(sum(row.amount for row in enriched_external_rows if row.is_reconciled), 2)
        selected_system_total = round(sum(row.amount for row in system_rows if row.is_reconciled), 2)
        return ReconciliationViewState(
            params=ReconciliationParams(account_id=account_id, date_from=date_from, date_to=date_to),
            accounts=self.list_accounts(),
            external_summary=external_summary,
            system_summary=system_summary,
            external_rows=enriched_external_rows,
            system_rows=system_rows,
            groups=groups,
            amount_difference=round(selected_external_total - selected_system_total, 2),
        )

    def reconcile_selection(
        self,
        *,
        account_id: int,
        date_from: str,
        date_to: str,
        system_transaction_ids: list[int],
        external_rows: tuple[ReconciliationExternalRow, ...],
    ) -> dict[str, Any]:
        """Return reconcile selection."""
        return self._db.reconciliation.reconcile(
            account_id=account_id,
            date_from=date_from,
            date_to=date_to,
            system_transaction_ids=system_transaction_ids,
            external_rows=[
                {
                    "date": row.date,
                    "reference": row.reference,
                    "description": row.description,
                    "amount": row.amount,
                    "external_item_key": row.external_item_key,
                }
                for row in external_rows
            ],
        )

    def clear_reconciliation_for_transactions(self, transaction_ids: list[int]) -> int:
        """Return clear reconciliation for transactions."""
        return self._db.reconciliation.clear_for_transactions(transaction_ids)
