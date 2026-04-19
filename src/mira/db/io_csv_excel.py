# SPDX-License-Identifier: GPL-3.0-or-later
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.

"""Module documentation."""

from __future__ import annotations


import csv
import logging
import re
import unicodedata
from pathlib import Path
from typing import TYPE_CHECKING, Any, Protocol, cast

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell as _OpenpyxlCell

from mira.db.money import MONEY_ZERO, MoneyLike, money_to_decimal

logger = logging.getLogger(__name__)

_CSV_IMPORT_ROW_ERRORS = (KeyError, TypeError, ValueError, OverflowError)
_TRANSACTION_SHEET_NAME = "Transactions"
_TRANSACTION_COLUMNS = (
    "id",
    "date",
    "type",
    "amount",
    "account_name",
    "category",
    "subcategory",
    "payment_method",
    "description",
    "note",
    "receipt_path",
    "tags",
)
_TRANSACTION_REQUIRED_COLUMNS = ("type", "amount")
_TRANSACTION_HEADER_ALIASES = {
    "id": ("id",),
    "date": ("date", "fecha"),
    "type": ("type", "tipo"),
    "amount": ("amount", "monto"),
    "account_name": ("account_name", "account", "cuenta", "nombre_cuenta"),
    "category": ("category", "categoria"),
    "subcategory": ("subcategory", "subcategoria"),
    "payment_method": ("payment_method", "metodo_pago", "medio_pago"),
    "description": ("description", "descripcion"),
    "note": ("note", "nota"),
    "receipt_path": ("receipt_path", "ruta_recibo", "ruta_comprobante", "comprobante"),
    "tags": ("tags", "etiquetas"),
}


def normalize_header(value: object) -> str:
    """Normalize a user-provided header into a canonical comparison key."""
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    folded = "".join(char for char in text if not unicodedata.combining(char)).casefold()
    separated = re.sub(r"[\s\-\/]+", "_", folded)
    cleaned = re.sub(r"[^0-9a-z_]", "", separated)
    return re.sub(r"_+", "_", cleaned).strip("_")


def _build_header_alias_lookup(aliases: dict[str, tuple[str, ...]]) -> dict[str, str]:
    """Build a normalized alias lookup keyed by header synonym."""
    lookup: dict[str, str] = {}
    for canonical, names in aliases.items():
        for name in names:
            lookup[normalize_header(name)] = canonical
    return lookup


_TRANSACTION_HEADER_ALIAS_LOOKUP = _build_header_alias_lookup(_TRANSACTION_HEADER_ALIASES)


def _unsupported_transaction_extension(filepath: str | Path) -> ValueError:
    """Build a consistent unsupported-extension error for transaction files."""
    suffix = Path(filepath).suffix or "<none>"
    return ValueError(f"Unsupported transaction file extension: {suffix}. Supported extensions are .csv and .xlsx.")


def _resolve_canonical_headers(
    headers: list[object],
    *,
    alias_lookup: dict[str, str],
    required_columns: tuple[str, ...],
    label: str,
) -> dict[str, int]:
    """Resolve source headers into canonical column names and detect collisions."""
    column_indexes: dict[str, int] = {}
    column_sources: dict[str, str] = {}
    for index, raw_header in enumerate(headers):
        normalized = normalize_header(raw_header)
        if not normalized:
            continue
        canonical = alias_lookup.get(normalized)
        if canonical is None:
            continue
        source_name = str(raw_header or "").strip() or normalized
        if canonical in column_indexes:
            previous = column_sources[canonical]
            raise ValueError(f"Ambiguous {label} headers for '{canonical}': {previous!r} and {source_name!r}.")
        column_indexes[canonical] = index
        column_sources[canonical] = source_name
    missing = [column for column in required_columns if column not in column_indexes]
    if missing:
        missing_list = ", ".join(missing)
        raise ValueError(f"Missing required {label} columns: {missing_list}.")
    return column_indexes


def _row_to_canonical_mapping(row: list[object], column_indexes: dict[str, int]) -> dict[str, object]:
    """Project a source row onto the canonical transaction schema."""
    canonical_row: dict[str, object] = {}
    for canonical, index in column_indexes.items():
        canonical_row[canonical] = row[index] if index < len(row) else None
    return canonical_row


def _transaction_export_rows(
    db: _DatabaseIOProtocol,
    *,
    tx_type: str | None = None,
    account_id: int | None = None,
    since_date: str | None = None,
    until_date: str | None = None,
    category: str | None = None,
    search: str | None = None,
) -> list[dict[str, object]]:
    """Collect transaction export rows including joined tag labels."""
    txs = db.get_transactions(
        limit=1_000_000,
        tx_type=tx_type,
        account_id=account_id,
        since_date=since_date,
        until_date=until_date,
        category=category,
        search=search,
    )
    tx_ids = [tx["id"] for tx in txs]
    tags_map = db.get_transactions_tags_bulk(tx_ids) if tx_ids else {}
    export_rows: list[dict[str, Any]] = []
    for tx in txs:
        row = {column: tx.get(column) for column in _TRANSACTION_COLUMNS}
        tag_list = tags_map.get(tx["id"], [])
        row["tags"] = ", ".join(t["name"] for t in tag_list)
        export_rows.append(row)
    return export_rows


def _import_transaction_row(db: _DatabaseIOProtocol, row: dict[str, object]) -> None:
    """Validate and insert a transaction row using canonical transaction fields."""
    account_name = str(row.get("account_name") or "General").strip() or "General"
    account = db.get_or_create_account(account_name)
    tx_type = str(row["type"] or "").strip().lower()
    if tx_type not in ("income", "expense"):
        raise ValueError(f"Invalid type: {tx_type}")
    amount = money_to_decimal(row["amount"])
    if amount is None:
        raise ValueError("Amount is required")
    if amount <= MONEY_ZERO:
        raise ValueError("Amount must be positive")
    tx = db.add_transaction(
        account_id=account["id"],
        tx_type=tx_type,
        amount=amount,
        description=str(row.get("description") or "").strip() or None,
        category=str(row.get("category") or "").strip() or None,
        subcategory=str(row.get("subcategory") or "").strip() or None,
        payment_method=str(row.get("payment_method") or "cash").strip() or "cash",
        tx_date=str(row.get("date") or "").strip() or None,
        note=str(row.get("note") or "").strip() or None,
        receipt_path=str(row.get("receipt_path") or "").strip() or None,
    )
    tags_str = str(row.get("tags") or "").strip()
    if not tags_str:
        return
    tag_ids: list[int] = []
    for tag_name in tags_str.split(","):
        normalized_tag_name = tag_name.strip()
        if not normalized_tag_name:
            continue
        tag = db.get_tag_by_name(normalized_tag_name)
        if tag is None:
            tag = db.add_tag(normalized_tag_name)
        tag_ids.append(int(tag["id"]))
    if tag_ids:
        db.set_transaction_tags(int(tx["id"]), tag_ids)


def _export_transactions_csv(
    db: _DatabaseIOProtocol,
    filepath: str,
    *,
    tx_type: str | None = None,
    account_id: int | None = None,
    since_date: str | None = None,
    until_date: str | None = None,
    category: str | None = None,
    search: str | None = None,
) -> int:
    """Export filtered transactions to CSV using canonical transaction columns."""
    rows = _transaction_export_rows(
        db,
        tx_type=tx_type,
        account_id=account_id,
        since_date=since_date,
        until_date=until_date,
        category=category,
        search=search,
    )
    with open(filepath, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(_TRANSACTION_COLUMNS), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(cast(Any, rows))
    return len(rows)


def _export_transactions_xlsx(
    db: _DatabaseIOProtocol,
    filepath: str | Path,
    *,
    tx_type: str | None = None,
    account_id: int | None = None,
    since_date: str | None = None,
    until_date: str | None = None,
    category: str | None = None,
    search: str | None = None,
) -> int:
    """Export filtered transactions to an XLSX workbook."""
    from openpyxl import Workbook

    rows = _transaction_export_rows(
        db,
        tx_type=tx_type,
        account_id=account_id,
        since_date=since_date,
        until_date=until_date,
        category=category,
        search=search,
    )
    target = Path(filepath).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _TRANSACTION_SHEET_NAME
    sheet.append(list(_TRANSACTION_COLUMNS))
    for row in rows:
        sheet.append([row.get(column) for column in _TRANSACTION_COLUMNS])
    workbook.save(target)
    return len(rows)


def _import_transactions_csv(db: _DatabaseIOProtocol, filepath: str) -> tuple[int, int]:
    """Import transactions from CSV using canonical header aliases."""
    imported = 0
    errors = 0
    with open(filepath, newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        header_row = next(reader, None)
        if header_row is None:
            raise ValueError("Transaction file is empty.")
        column_indexes = _resolve_canonical_headers(
            list(header_row),
            alias_lookup=_TRANSACTION_HEADER_ALIAS_LOOKUP,
            required_columns=_TRANSACTION_REQUIRED_COLUMNS,
            label="transaction",
        )
        for row in reader:
            try:
                _import_transaction_row(db, _row_to_canonical_mapping(list(row), column_indexes))
                imported += 1
            except _CSV_IMPORT_ROW_ERRORS as exc:
                logger.warning("CSV import error on row %r: %s", row, exc)
                errors += 1
    return imported, errors


def _import_transactions_xlsx(db: _DatabaseIOProtocol, filepath: str) -> tuple[int, int]:
    """Import transactions from XLSX using canonical header aliases."""
    from openpyxl import load_workbook

    imported = 0
    errors = 0
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    try:
        worksheet = cast(
            Any,
            workbook[_TRANSACTION_SHEET_NAME] if _TRANSACTION_SHEET_NAME in workbook.sheetnames else workbook.active,
        )
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError("Transaction file is empty.")
        column_indexes = _resolve_canonical_headers(
            list(header_row),
            alias_lookup=_TRANSACTION_HEADER_ALIAS_LOOKUP,
            required_columns=_TRANSACTION_REQUIRED_COLUMNS,
            label="transaction",
        )
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            try:
                _import_transaction_row(db, _row_to_canonical_mapping(list(row), column_indexes))
                imported += 1
            except _CSV_IMPORT_ROW_ERRORS as exc:
                logger.warning("XLSX import error on row %r: %s", row, exc)
                errors += 1
    finally:
        workbook.close()
    return imported, errors


def import_transactions_file(db: _DatabaseIOProtocol, filepath: str) -> tuple[int, int]:
    """Import transactions from CSV or XLSX based on the file extension."""
    suffix = Path(filepath).suffix.casefold()
    if suffix == ".csv":
        return _import_transactions_csv(db, filepath)
    if suffix == ".xlsx":
        return _import_transactions_xlsx(db, filepath)
    raise _unsupported_transaction_extension(filepath)


def export_transactions_file(
    db: _DatabaseIOProtocol,
    filepath: str,
    *,
    tx_type: str | None = None,
    account_id: int | None = None,
    since_date: str | None = None,
    until_date: str | None = None,
    category: str | None = None,
    search: str | None = None,
) -> int:
    """Export transactions to CSV or XLSX based on the file extension."""
    suffix = Path(filepath).suffix.casefold()
    if suffix == ".csv":
        return _export_transactions_csv(
            db,
            filepath,
            tx_type=tx_type,
            account_id=account_id,
            since_date=since_date,
            until_date=until_date,
            category=category,
            search=search,
        )
    if suffix == ".xlsx":
        return _export_transactions_xlsx(
            db,
            filepath,
            tx_type=tx_type,
            account_id=account_id,
            since_date=since_date,
            until_date=until_date,
            category=category,
            search=search,
        )
    raise _unsupported_transaction_extension(filepath)


class _DatabaseIOProtocol(Protocol):
    """Represent the _DatabaseIOProtocol class."""

    def get_transactions(
        self,
        *,
        limit: int = 50,
        tx_type: str | None = None,
        account_id: int | None = None,
        since_date: str | None = None,
        until_date: str | None = None,
        category: str | None = None,
        payment_method: str | None = None,
        min_amount: MoneyLike | None = None,
        max_amount: MoneyLike | None = None,
        search: str | None = None,
        tag_id: int | None = None,
        include_children: bool = False,
    ) -> list[dict]:
        """Return get transactions."""
        ...

    def get_transactions_tags_bulk(self, transaction_ids: list[int]) -> dict[int, list[dict]]:
        """Return get transactions tags bulk."""
        ...

    def get_budget_comparison(self, budget_id: int, granularity: str = "quarterly") -> dict[str, object]:
        """Return get budget comparison."""
        ...

    def get_or_create_account(self, name: str) -> dict:
        """Return get or create account."""
        ...

    def get_setting(self, key: str) -> str | None:
        """Return get setting."""
        ...

    def add_transaction(
        self,
        *,
        account_id: int,
        tx_type: str,
        amount: MoneyLike,
        description: str | None = None,
        category: str | None = None,
        subcategory: str | None = None,
        payment_method: str = "cash",
        receipt_path: str | None = None,
        tx_date: str | None = None,
        note: str | None = None,
        to_account_id: int | None = None,
        is_transfer: int = 0,
        exchange_rate: float | None = None,
        converted_amount: MoneyLike | None = None,
        category_id: int | None = None,
        source: str | None = None,
    ) -> dict:
        """Return add transaction."""
        ...

    def get_or_create_tag(self, name: str) -> dict:
        """Return get or create tag."""
        ...

    def get_tag_by_name(self, name: str) -> dict | None:
        """Return get tag by name."""
        ...

    def add_tag(self, name: str, color: str = "#888888", icon: str = "") -> dict:
        """Return add tag."""
        ...

    def add_transaction_tag(self, transaction_id: int, tag_id: int) -> None:
        """Return add transaction tag."""
        ...

    def set_transaction_tags(self, transaction_id: int, tag_ids: list[int]) -> None:
        """Return set transaction tags."""
        ...


def export_transactions_csv(
    db: _DatabaseIOProtocol,
    filepath: str,
    *,
    tx_type: str | None = None,
    account_id: int | None = None,
    since_date: str | None = None,
    until_date: str | None = None,
    category: str | None = None,
    search: str | None = None,
) -> int:
    """Export filtered transactions to a CSV file. Returns row count."""
    return _export_transactions_csv(
        db,
        filepath,
        tx_type=tx_type,
        account_id=account_id,
        since_date=since_date,
        until_date=until_date,
        category=category,
        search=search,
    )


def _variance_signal(section: str, variance: float) -> str:
    """Return variance signal."""
    if abs(variance) < 0.005:
        return "neutral"
    if section == "income":
        return "positive" if variance > 0 else "negative"
    if section == "expense":
        return "negative" if variance > 0 else "positive"
    return "positive" if variance > 0 else "negative"


def _wc(cell: Any) -> "_OpenpyxlCell":
    """Cast an openpyxl cell to a writable Cell (stubs return Union[Cell, ReadOnlyCell, MergedCell])."""
    return cast("_OpenpyxlCell", cell)


def export_budget_comparison_excel(
    db: _DatabaseIOProtocol,
    filepath: str | Path,
    budget_id: int,
    *,
    granularity: str = "quarterly",
) -> int:
    """Export the real-vs-budget comparison to an Excel workbook. Returns exported row count."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    comparison = db.get_budget_comparison(budget_id, granularity=granularity)
    budget = cast(dict[str, Any], comparison["budget"])
    periods = cast(list[dict[str, Any]], comparison["periods"])
    rows = cast(list[dict[str, Any]], comparison["rows"])
    totals = cast(dict[str, Any], comparison["totals"])
    excluded_transactions = int(cast(Any, comparison["excluded_transactions"]))

    headers = ["Categoría"]
    for period in periods:
        label = str(period["label"])
        headers.extend([f"{label} Real", f"{label} PPTO", f"{label} Var"])
    headers.extend(["Año Real", "Año PPTO", "Año Var"])

    incomes = [row for row in rows if row["type"] == "income"]
    expenses = [row for row in rows if row["type"] == "expense"]
    structure = [
        ("section", "Ingresos"),
        *[("category", row) for row in incomes],
        ("total_income", None),
        ("section", "Gastos"),
        *[("category", row) for row in expenses],
        ("total_expense", None),
        ("balance", None),
    ]

    target = Path(filepath).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Real vs PPTO"
    sheet.freeze_panes = cast(Any, "A9")

    title = f"Reporte Real vs Presupuesto | {budget['code']} | {budget['year']} | {budget['currency']}"
    sheet.cell(1, 1, title)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    meta_rows: list[tuple[str, str | int | None]] = [
        ("Presupuesto", str(budget["code"])),
        ("Año", int(budget["year"])),
        ("Moneda", str(budget["currency"])),
        (
            "Vista",
            {
                "monthly": "Mensual",
                "quarterly": "Trimestral",
                "semiannual": "Semestral",
                "annual": "Anual",
            }.get(granularity, granularity),
        ),
        ("Transacciones excluidas", excluded_transactions),
    ]
    for row_idx, (label, value) in enumerate(meta_rows, start=2):
        sheet.cell(row_idx, 1, label)
        sheet.cell(row_idx, 2, value)

    title_fill = PatternFill("solid", fgColor="263140")
    header_fill = PatternFill("solid", fgColor="2F3640")
    section_fill = PatternFill("solid", fgColor="263140")
    income_total_fill = PatternFill("solid", fgColor="21332E")
    expense_total_fill = PatternFill("solid", fgColor="352826")
    balance_fill = PatternFill("solid", fgColor="223042")
    positive_fill = PatternFill("solid", fgColor="173B36")
    negative_fill = PatternFill("solid", fgColor="472624")
    neutral_fill = PatternFill("solid", fgColor="30363D")

    title_font = Font(color="F0F3F7", bold=True, size=14)
    white_font = Font(color="FFFFFF", bold=True)
    section_font = Font(color="D6DEE8", bold=True)
    positive_font = Font(color="7FE7D2", bold=True)
    negative_font = Font(color="FFB0A3", bold=True)
    neutral_font = Font(color="C4CCD5", bold=True)

    thin_border = Border(
        left=Side(style="thin", color="4A5563"),
        right=Side(style="thin", color="4A5563"),
        top=Side(style="thin", color="4A5563"),
        bottom=Side(style="thin", color="4A5563"),
    )

    _wc(sheet["A1"]).fill = title_fill
    _wc(sheet["A1"]).font = title_font
    _wc(sheet["A1"]).alignment = Alignment(horizontal="left", vertical="center")

    for row_idx in range(2, 7):
        _wc(sheet.cell(row_idx, 1)).font = Font(bold=True)

    header_row = 8
    for col_idx, header in enumerate(headers, start=1):
        cell = _wc(sheet.cell(header_row, col_idx, header))
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx = 9
    for kind, payload in structure:
        if kind == "section":
            for col_idx in range(1, len(headers) + 1):
                cell = _wc(sheet.cell(row_idx, col_idx))
                cell.fill = section_fill
                cell.border = thin_border
            section_cell = _wc(sheet.cell(row_idx, 1, str(payload)))
            section_cell.font = section_font
            row_idx += 1
            continue

        if kind == "category":
            category = cast(dict[str, Any], payload)
            section = str(category["type"])
            sheet.cell(row_idx, 1, str(category["name"]))
            col_idx = 2
            for period in cast(list[dict[str, Any]], category["periods"]):
                sheet.cell(row_idx, col_idx, float(period["real"]))
                sheet.cell(row_idx, col_idx + 1, float(period["budget"]))
                variance_cell = _wc(sheet.cell(row_idx, col_idx + 2, float(period["variance"])))
                signal = _variance_signal(section, float(period["variance"]))
                variance_cell.fill = {
                    "positive": positive_fill,
                    "negative": negative_fill,
                    "neutral": neutral_fill,
                }[signal]
                variance_cell.font = {
                    "positive": positive_font,
                    "negative": negative_font,
                    "neutral": neutral_font,
                }[signal]
                col_idx += 3
            sheet.cell(row_idx, col_idx, float(category["annual_real"]))
            sheet.cell(row_idx, col_idx + 1, float(category["annual_budget"]))
            annual_variance_cell = _wc(sheet.cell(row_idx, col_idx + 2, float(category["annual_variance"])))
            annual_signal = _variance_signal(section, float(category["annual_variance"]))
            annual_variance_cell.fill = {
                "positive": positive_fill,
                "negative": negative_fill,
                "neutral": neutral_fill,
            }[annual_signal]
            annual_variance_cell.font = {
                "positive": positive_font,
                "negative": negative_font,
                "neutral": neutral_font,
            }[annual_signal]
        else:
            total_key = "income" if kind == "total_income" else "expense" if kind == "total_expense" else "balance"
            row_fill = (
                income_total_fill
                if kind == "total_income"
                else expense_total_fill if kind == "total_expense" else balance_fill
            )
            title_text = (
                "Subtotal ingresos"
                if kind == "total_income"
                else "Subtotal gastos" if kind == "total_expense" else "Balance"
            )
            period_totals = cast(list[dict[str, Any]], totals[total_key])
            annual_totals = cast(dict[str, Any], totals[f"{total_key}_annual"])
            title_cell = _wc(sheet.cell(row_idx, 1, title_text))
            title_cell.font = section_font
            col_idx = 2
            for period in period_totals:
                sheet.cell(row_idx, col_idx, float(period["real"]))
                sheet.cell(row_idx, col_idx + 1, float(period["budget"]))
                variance_cell = _wc(sheet.cell(row_idx, col_idx + 2, float(period["variance"])))
                signal = _variance_signal(total_key, float(period["variance"]))
                variance_cell.fill = {
                    "positive": positive_fill,
                    "negative": negative_fill,
                    "neutral": neutral_fill,
                }[signal]
                variance_cell.font = {
                    "positive": positive_font,
                    "negative": negative_font,
                    "neutral": neutral_font,
                }[signal]
                col_idx += 3
            sheet.cell(row_idx, col_idx, float(annual_totals["real"]))
            sheet.cell(row_idx, col_idx + 1, float(annual_totals["budget"]))
            annual_variance_cell = _wc(sheet.cell(row_idx, col_idx + 2, float(annual_totals["variance"])))
            annual_signal = _variance_signal(total_key, float(annual_totals["variance"]))
            annual_variance_cell.fill = {
                "positive": positive_fill,
                "negative": negative_fill,
                "neutral": neutral_fill,
            }[annual_signal]
            annual_variance_cell.font = {
                "positive": positive_font,
                "negative": negative_font,
                "neutral": neutral_font,
            }[annual_signal]
            for col_idx in range(1, len(headers) + 1):
                cell = _wc(sheet.cell(row_idx, col_idx))
                if not cell.fill.fill_type:
                    cell.fill = row_fill
                if col_idx != 1 and not cell.font.bold:
                    cell.font = Font(color="F0F3F7", bold=True)

        for col_idx in range(1, len(headers) + 1):
            cell = _wc(sheet.cell(row_idx, col_idx))
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00"
        row_idx += 1

    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row_idx - 1}"
    for col_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 24 if col_idx == 1 else 14

    workbook.save(target)
    return len(structure)


def import_transactions_csv(db: _DatabaseIOProtocol, filepath: str) -> tuple[int, int]:
    """Import transactions from a CSV file. Returns (imported, errors)."""
    return _import_transactions_csv(db, filepath)
