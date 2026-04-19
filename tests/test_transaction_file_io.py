# SPDX-License-Identifier: GPL-3.0-or-later
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.

from __future__ import annotations

import csv
from collections.abc import Sequence
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook

from mira.db.database import Database
from tests.db_inspection import fetch_all_dicts


@pytest.fixture
def db(tmp_path):
    database = Database(path=tmp_path / "test.db")
    database.connect()
    yield database
    database.close()


def _write_csv(path, headers: Sequence[str], rows: Sequence[dict[str, object]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(
    path,
    headers: Sequence[str],
    rows: Sequence[dict[str, object]],
    *,
    sheet_name: str = "Transactions",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    workbook.save(path)


def _seed_transactions(db: Database) -> None:
    account = db.account.get_or_create("General")
    db.transaction.create(
        account_id=account["id"],
        tx_type="income",
        amount=1000.10,
        description="Salary",
        category="Income",
        tx_date="2026-02-05",
    )
    db.transaction.create(
        account_id=account["id"],
        tx_type="expense",
        amount=300.20,
        description="Rent",
        category="Housing",
        tx_date="2026-02-10",
    )


def test_export_transactions_file_xlsx_writes_headers_and_rows(db, tmp_path) -> None:
    _seed_transactions(db)

    filepath = tmp_path / "transactions.xlsx"
    count = db.io.export_transactions_file(str(filepath))

    workbook = load_workbook(filepath, data_only=True)
    try:
        sheet = cast(Any, workbook["Transactions"])
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert count == 2
    assert rows[0] == (
        "id",
        "date",
        "type",
        "amount",
        "account_name",
        "category",
        "subcategory",
        "payment_method",
        "description",
        "note",
        "receipt_path",
        "tags",
    )
    exported_types = {row[2] for row in rows[1:]}
    assert exported_types == {"income", "expense"}


def test_export_transactions_file_xlsx_respects_type_filter(db, tmp_path) -> None:
    _seed_transactions(db)

    filepath = tmp_path / "expenses.xlsx"
    count = db.io.export_transactions_file(str(filepath), tx_type="expense")

    workbook = load_workbook(filepath, data_only=True)
    try:
        sheet = cast(Any, workbook["Transactions"])
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert count == 1
    assert len(rows) == 2
    assert rows[1][2] == "expense"


def test_import_transactions_file_xlsx_reads_valid_rows(db, tmp_path) -> None:
    filepath = tmp_path / "import.xlsx"
    _write_xlsx(
        filepath,
        ["date", "type", "amount", "account_name", "description", "tags"],
        [
            {
                "date": "2026-03-01",
                "type": "income",
                "amount": "2500.00",
                "account_name": "General",
                "description": "Salary March",
                "tags": "job, payroll",
            },
            {
                "date": "2026-03-05",
                "type": "expense",
                "amount": "90.50",
                "account_name": "General",
                "description": "Internet",
                "tags": "home",
            },
        ],
    )

    imported, errors = db.io.import_transactions_file(str(filepath))

    assert imported == 2
    assert errors == 0
    assert {float(item["amount"]) for item in db.transaction.list()} == {2500.0, 90.5}
    assert {tag["name"] for tag in db.tag.list()} >= {"job", "payroll", "home"}


def test_import_transactions_file_xlsx_skips_invalid_rows_and_counts_errors(db, tmp_path) -> None:
    filepath = tmp_path / "invalid_rows.xlsx"
    _write_xlsx(
        filepath,
        ["date", "type", "amount", "description"],
        [
            {"date": "2026-03-01", "type": "income", "amount": "500.0", "description": "OK"},
            {"date": "2026-03-02", "type": "transfer", "amount": "100.0", "description": "Bad type"},
            {"date": "2026-03-03", "type": "expense", "amount": "-20", "description": "Negative"},
        ],
    )

    imported, errors = db.io.import_transactions_file(str(filepath))

    assert imported == 1
    assert errors == 2


def test_export_import_xlsx_round_trip_preserves_fractional_cents_exactly(db, tmp_path) -> None:
    _seed_transactions(db)
    filepath = tmp_path / "round_trip.xlsx"
    db.io.export_transactions_file(str(filepath))

    db.close()
    fresh_db = Database(path=tmp_path / "fresh.db")
    fresh_db.connect()
    try:
        imported, errors = fresh_db.io.import_transactions_file(str(filepath))
        assert (imported, errors) == (2, 0)
        stored = fetch_all_dicts(fresh_db, "SELECT amount_cents FROM transactions ORDER BY amount_cents")
        assert [int(row["amount_cents"]) for row in stored] == [30020, 100010]
    finally:
        fresh_db.close()


def test_export_transactions_file_dispatches_by_extension(db, tmp_path) -> None:
    _seed_transactions(db)

    csv_path = tmp_path / "transactions.csv"
    xlsx_path = tmp_path / "transactions.xlsx"

    assert db.io.export_transactions_file(str(csv_path)) == 2
    assert db.io.export_transactions_file(str(xlsx_path)) == 2

    with open(csv_path, encoding="utf-8") as fh:
        assert fh.readline().strip().startswith("id,date,type,amount")

    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        assert workbook.sheetnames == ["Transactions"]
    finally:
        workbook.close()


@pytest.mark.parametrize("extension", [".xls", ".txt"])
def test_transaction_file_apis_reject_xls_and_other_extensions(db, tmp_path, extension: str) -> None:
    filepath = tmp_path / f"transactions{extension}"

    with pytest.raises(ValueError, match="Unsupported transaction file extension"):
        db.io.import_transactions_file(str(filepath))

    with pytest.raises(ValueError, match="Unsupported transaction file extension"):
        db.io.export_transactions_file(str(filepath))


@pytest.mark.parametrize(
    ("extension", "headers"),
    [
        (".csv", ["date", "type", "amount", "account_name", "description"]),
        (".csv", ["fecha", "tipo", "monto", "cuenta", "descripcion"]),
        (".csv", ["Fecha", "type", "Monto", "account", "description"]),
        (".xlsx", ["date", "type", "amount", "account_name", "description"]),
        (".xlsx", ["fecha", "tipo", "monto", "cuenta", "descripcion"]),
        (".xlsx", ["Fecha", "type", "Monto", "account", "description"]),
    ],
)
def test_transaction_import_accepts_bilingual_headers(db, tmp_path, extension: str, headers: list[str]) -> None:
    filepath = tmp_path / f"headers{extension}"
    rows = [
        {
            headers[0]: "2026-04-01",
            headers[1]: "income",
            headers[2]: "1200.50",
            headers[3]: "Caja",
            headers[4]: "Ingreso por prueba",
        }
    ]
    if extension == ".csv":
        _write_csv(filepath, headers, rows)
    else:
        _write_xlsx(filepath, headers, rows)

    imported, errors = db.io.import_transactions_file(str(filepath))

    assert (imported, errors) == (1, 0)
    tx = db.transaction.list(limit=5)[0]
    assert tx["description"] == "Ingreso por prueba"
    assert float(tx["amount"]) == pytest.approx(1200.50)


@pytest.mark.parametrize("extension", [".csv", ".xlsx"])
def test_transaction_import_rejects_alias_collisions(db, tmp_path, extension: str) -> None:
    filepath = tmp_path / f"collision{extension}"
    headers = ["type", "tipo", "amount"]
    rows = [{"type": "income", "tipo": "expense", "amount": "99.00"}]
    if extension == ".csv":
        _write_csv(filepath, headers, rows)
    else:
        _write_xlsx(filepath, headers, rows)

    with pytest.raises(ValueError, match="Ambiguous transaction headers"):
        db.io.import_transactions_file(str(filepath))


@pytest.mark.parametrize("extension", [".csv", ".xlsx"])
def test_transaction_import_requires_type_and_amount_headers(db, tmp_path, extension: str) -> None:
    filepath = tmp_path / f"missing{extension}"
    headers = ["date", "description"]
    rows = [{"date": "2026-04-01", "description": "Missing required columns"}]
    if extension == ".csv":
        _write_csv(filepath, headers, rows)
    else:
        _write_xlsx(filepath, headers, rows)

    with pytest.raises(ValueError, match="Missing required transaction columns"):
        db.io.import_transactions_file(str(filepath))
